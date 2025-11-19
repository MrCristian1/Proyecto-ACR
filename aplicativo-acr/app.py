import smtplib
from email.message import EmailMessage
import time

def enviar_acr_smtp_env(archivo_bytes, nombre_archivo):
    import os
    from dotenv import load_dotenv
    load_dotenv()
    
    # Rate limiting: verificar último envío
    ultimo_envio = st.session_state.get('ultimo_envio_email', 0)
    tiempo_actual = time.time()
    TIEMPO_MINIMO_ENTRE_ENVIOS = 300  # 5 minutos en segundos
    
    if tiempo_actual - ultimo_envio < TIEMPO_MINIMO_ENTRE_ENVIOS:
        tiempo_restante = int((TIEMPO_MINIMO_ENTRE_ENVIOS - (tiempo_actual - ultimo_envio)) / 60)
        st.error(f"⚠️ Debes esperar {tiempo_restante + 1} minuto(s) antes de enviar otro correo.")
        return False
    
    # Validar límite diario
    hoy = datetime.now().date().isoformat()
    if st.session_state.get('fecha_ultimo_envio') != hoy:
        st.session_state['emails_enviados_hoy'] = 0
        st.session_state['fecha_ultimo_envio'] = hoy
    
    LIMITE_DIARIO = 10  # Máximo 10 correos por día
    if st.session_state.get('emails_enviados_hoy', 0) >= LIMITE_DIARIO:
        st.error(f"⚠️ Se alcanzó el límite de {LIMITE_DIARIO} correos por día.")
        return False
    
    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587
    SMTP_USER = os.getenv("SMTP_USER")
    SMTP_PASS = os.getenv("SMTP_PASS")
    DESTINATARIO = os.getenv("DESTINATARIO")

    if not SMTP_USER or not SMTP_PASS or not DESTINATARIO:
        st.error("Faltan variables SMTP_USER, SMTP_PASS o DESTINATARIO en el archivo .env")
        return False
    
    # Validar que el destinatario sea del dominio permitido
    DOMINIO_PERMITIDO = "solutionsandpayroll.com"
    if not DESTINATARIO.endswith(f"@{DOMINIO_PERMITIDO}"):
        st.error(f"⚠️ Solo se permite enviar correos al dominio {DOMINIO_PERMITIDO}")
        return False

    msg = EmailMessage()
    # Extraer consecutivo del nombre del archivo (asume formato '026 ACCIONES ...')
    consecutivo = nombre_archivo.split()[0] if nombre_archivo else ""
    msg['Subject'] = f"ACR {consecutivo} generado - Solutions & Payroll"
    msg['From'] = f"Automatizacion ACR <{SMTP_USER}>"
    msg['To'] = DESTINATARIO
    msg.set_content(f"A continuación encontrarás el reporte ACR '{nombre_archivo}' generado por el sistema.")

    # Adjuntar el archivo Excel
    msg.add_attachment(
        archivo_bytes,
        maintype='application',
        subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=nombre_archivo
    )

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.starttls()
            smtp.login(SMTP_USER, SMTP_PASS)
            smtp.send_message(msg)
        
        # Actualizar contadores
        st.session_state['ultimo_envio_email'] = tiempo_actual
        st.session_state['emails_enviados_hoy'] = st.session_state.get('emails_enviados_hoy', 0) + 1
        
        return True
    except Exception as e:
        st.error(f"Error al enviar el correo: {e}")
        return False

import streamlit as st
import pandas as pd
import openpyxl
import requests
import json
from datetime import datetime, date
from io import BytesIO
import os
from dotenv import load_dotenv
import base64
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Cargar variables de entorno
load_dotenv()

# Configuración de la página
st.set_page_config(
    page_title="Solutions & Payroll - ACR Manager",
    page_icon="syp logo.png",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para interfaz empresarial
def load_custom_css():
    st.markdown("""
    <style>
        /* Estilos globales */
        .main {
            background-color: #f5f7fa;
        }
        
        /* Header principal */
        /* Header principal mejorado */
        .header-container {
            background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #334155 100%);
            padding: 3rem 2rem;
            border-radius: 16px;
            margin-bottom: 2.5rem;
            box-shadow: 
                0 12px 40px 0 rgba(30, 58, 138, 0.4),
                0 4px 12px 0 rgba(0, 0, 0, 0.3),
                inset 0 1px 0 rgba(255, 255, 255, 0.1);
            border-top: 8px solid #2563eb;
            border-image: none;
            position: relative;
            overflow: hidden;
            transition: all 0.3s ease;
        }

        /* Efecto de brillo sutil */
        .header-container::before {
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(
                90deg,
                transparent,
                rgba(255, 255, 255, 0.05),
                transparent
            );
            transition: left 0.6s ease;
        }

        .header-container:hover::before {
            left: 100%;
        }

        /* Efecto de partículas sutiles */
        .header-container::after {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: 
                radial-gradient(circle at 20% 80%, rgba(59, 130, 246, 0.1) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, rgba(162, 28, 175, 0.1) 0%, transparent 50%),
                radial-gradient(circle at 40% 40%, rgba(245, 158, 66, 0.05) 0%, transparent 50%);
            pointer-events: none;
        }

        /* Efecto hover mejorado */
        .header-container:hover {
            transform: translateY(-2px);
            box-shadow: 
                0 16px 50px 0 rgba(30, 58, 138, 0.5),
                0 6px 16px 0 rgba(0, 0, 0, 0.35),
                inset 0 1px 0 rgba(255, 255, 255, 0.15);
        }

        /* Para el contenido dentro del header */
        .header-content {
            position: relative;
            z-index: 2;
            text-align: left;
        }

        /* Títulos dentro del header */
        .header-title {
            color: #fff;
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 1rem;
            background: none;
            -webkit-background-clip: initial;
            -webkit-text-fill-color: initial;
            background-clip: initial;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }

        .header-subtitle {
            color: #cbd5e1;
            font-size: 1.2rem;
            font-weight: 400;
            line-height: 1.6;
            max-width: 600px;
            margin: 0 auto;
        }
        
        .company-logo {
            font-size: 3rem;
            display: inline-block;
            margin-right: 1rem;
            vertical-align: middle;
        }
        
        .company-name {
            color: white;
            font-size: 2.5rem;
            font-weight: 700;
            display: inline-block;
            vertical-align: middle;
            margin: 0;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.2);
        }
        
        .company-subtitle {
            color: #e0e7ff;
            font-size: 1.1rem;
            margin-top: 0.5rem;
            font-weight: 300;
        }
        
        /* Sidebar mejorado y moderno */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
            border-right: 1px solid rgba(255, 255, 255, 0.1);
        }
        
        [data-testid="stSidebar"] .element-container {
            color: #e2e8f0;
        }
        
        /* Títulos del sidebar */
        [data-testid="stSidebar"] h3 {
            color: #ffffff;
            font-weight: 600;
            font-size: 1.1rem;
            margin-bottom: 1rem;
            padding: 0.5rem 0;
            border-bottom: 2px solid #2563eb;
        }
        
        /* Separadores del sidebar */
        [data-testid="stSidebar"] hr {
            border: none;
            height: 1px;
            background: rgba(255, 255, 255, 0.1);
            margin: 1.5rem 0;
        }
        
        /* Radio buttons del sidebar */
        [data-testid="stSidebar"] .stRadio > div {
            background: rgba(255, 255, 255, 0.03);
            border-radius: 8px;
            padding: 0.5rem;
            border: 1px solid rgba(255, 255, 255, 0.05);
        }
        
        /* Info boxes del sidebar */
        [data-testid="stSidebar"] .stAlert {
            background: rgba(37, 99, 235, 0.1);
            border: 1px solid rgba(37, 99, 235, 0.3);
            border-radius: 8px;
            color: #cbd5e1;
        }
        
        /* Expanders del sidebar */
        [data-testid="stSidebar"] .streamlit-expanderHeader {
            background: rgba(255, 255, 255, 0.05);
            border-radius: 6px;
            color: #f1f5f9;
            font-weight: 500;
        }
        
        [data-testid="stSidebar"] .streamlit-expanderContent {
            background: rgba(255, 255, 255, 0.02);
            border-radius: 0 0 6px 6px;
            color: #cbd5e1;
        }
        
        /* Tarjetas de formulario */
        .stForm {
            padding: 2rem;
            border-radius: 10px;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        }
        
        /* Botones minimalistas */
        .stButton > button {
            background: rgba(255, 255, 255, 0.05);
            color: #e2e8f0;
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 6px;
            padding: 0.6rem 1.5rem;
            font-weight: 500;
            font-size: 0.9rem;
            transition: all 0.2s ease;
            box-shadow: none;
            backdrop-filter: blur(10px);
        }
        
        .stButton > button:hover {
            background: rgba(255, 255, 255, 0.1);
            border-color: rgba(255, 255, 255, 0.2);
            color: #ffffff;
            transform: translateY(-1px);
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.2);
        }
        
        .stButton > button:active {
            transform: translateY(0px);
            background: rgba(255, 255, 255, 0.15);
        }
        
        /* Botones de agregar/quitar con iconos azules */
        .stButton > button[aria-label*="Agregar"], 
        .stButton > button:has(div:contains("Agregar")),
        .stButton > button:has(div:contains("➕")) {
            color: #3b82f6;
        }
        
        .stButton > button[aria-label*="Quitar"], 
        .stButton > button:has(div:contains("Quitar")),
        .stButton > button:has(div:contains("➖")) {
            color: #3b82f6;
        }
        
        /* Info boxes */
        .stAlert {
            border-radius: 8px;
            border-left: 4px solid #3b82f6;
        }
        
        /* Secciones */
        .section-header {
            background: linear-gradient(90deg, #3b82f6 0%, transparent 100%);
            padding: 1rem;
            border-radius: 8px;
            margin: 1.5rem 0 1rem 0;
        }
        
        .section-header h3 {
            color: white;
            margin: 0;
            font-weight: 600;
        }
        
        /* Divisor personalizado */
        hr {
            margin: 2rem 0;
            border: none;
            height: 2px;
            background: linear-gradient(90deg, transparent, #3b82f6, transparent);
        }
        
        /* Input fields */
        .stTextInput > div > div > input,
        .stTextArea > div > div > textarea,
        .stDateInput > div > div > input {
            border-radius: 8px;
            transition: border-color 0.3s ease;
        }
        
        .stTextInput > div > div > input:focus,
        .stTextArea > div > div > textarea:focus,
        .stDateInput > div > div > input:focus {
            border-color: #3b82f6;
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
        }
        
        /* Tarjeta de estadísticas */
        .stat-card {
            padding: 1.5rem;
            border-radius: 10px;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
            border-left: 4px solid #3b82f6;
        }
    </style>
    """, unsafe_allow_html=True)

def render_header():
    """Renderiza el encabezado empresarial"""
    # Convertir la imagen a base64
    logo_path = "syp logo.png"
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as image_file:
            encoded = base64.b64encode(image_file.read()).decode()
        img_html = f'<img src="data:image/png;base64,{encoded}" width="64" style="margin-right: 16px;vertical-align:middle;"/>'
    else:
        img_html = ""
    st.markdown(f"""
    <div class="header-container">
        <div class="header-content">
            <div style="display: flex; align-items: center; gap: 16px;">
                {img_html}
                <h1 class="header-title" style="margin: 0;">Solutions & Payroll</h1>
            </div>
            <p class="header-subtitle">Sistema de Gestión de Análisis de Causa Raíz (ACR)</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

def main():
    load_custom_css()
    render_header()
    
    # Sidebar mejorado
    with st.sidebar:
        st.markdown("### 📋 Panel de Navegación")
        st.markdown("---")
        
        opcion = st.radio(
            "Selecciona una opción:",
            ["🆕 Crear Nueva ACR", # "📤 Cargar ACR Existente", 
             "📊 Información del Sistema"],
            label_visibility="collapsed"
        )
        
        st.markdown("---")
        st.markdown("### 💡 Ayuda Rápida")
        with st.expander("¿Qué es un ACR?"):
            st.write("El Análisis de Causa Raíz (ACR) es una metodología para identificar las causas fundamentales de un problema.")
        with st.expander("¿Cómo usar el sistema?"):
            st.write("1. Completa el formulario con los datos del problema\n2. Genera análisis con IA o manualmente\n3. Descarga el reporte en Excel")
        st.markdown("---")
        st.markdown("### 📌 Información")
        st.markdown('<span style="color: #fff; font-weight: 500;">Versión: 2.0<br>Última actualización: Oct 2025</span>', unsafe_allow_html=True)
    
    # Contenido principal
    if opcion == "🆕 Crear Nueva ACR":
        crear_nueva_acr()
    # elif opcion == "📤 Cargar ACR Existente":
    #     cargar_acr_existente()
    elif opcion == "📊 Información del Sistema":
        mostrar_informacion_sistema()


def conectar_google_sheets():
    """Conecta con Google Sheets usando credenciales"""
    try:
        # Intentar cargar credenciales desde Streamlit Secrets (producción)
        try:
            if 'gcp_service_account' in st.secrets:
                credentials = ServiceAccountCredentials.from_json_keyfile_dict(
                    st.secrets["gcp_service_account"],
                    ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
                )
                sheet_name = st.secrets.get("GOOGLE_SHEET_NAME", "ACR_Consecutivos")
                client = gspread.authorize(credentials)
                sheet = client.open(sheet_name).sheet1
                return sheet
        except (FileNotFoundError, AttributeError):
            # No hay secrets.toml, continuar con archivo local
            pass
        
        # Desarrollo local: usar archivo JSON
        load_dotenv()
        credentials_file = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "google_credentials.json")
        sheet_name = os.getenv("GOOGLE_SHEET_NAME", "ACR_Consecutivos")
        
        if not os.path.exists(credentials_file):
            print(f"Archivo de credenciales no encontrado: {credentials_file}")
            return None
        
        credentials = ServiceAccountCredentials.from_json_keyfile_name(
            credentials_file,
            ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        )
        
        client = gspread.authorize(credentials)
        sheet = client.open(sheet_name).sheet1
        return sheet
    except Exception as e:
        print(f"Error conectando con Google Sheets: {e}")
        return None

def leer_consecutivo():
    """Lee el consecutivo desde Google Sheets"""
    try:
        sheet = conectar_google_sheets()
        if sheet:
            # Leer valor de la celda A2
            valor = sheet.acell('A2').value
            if valor and valor.isdigit():
                return int(valor)
        # Si falla Google Sheets, usar archivo local como fallback
        if os.path.exists("consecutivo.txt"):
            with open("consecutivo.txt", "r") as f:
                valor = f.read().strip()
                if valor.isdigit():
                    return int(valor)
    except Exception as e:
        print(f"Error leyendo consecutivo: {e}")
    return 26  # valor por defecto inicial

def consecutivo_formateado(numero):
    return f"{int(numero):03d}"

def guardar_consecutivo(nuevo_valor):
    """Guarda el consecutivo en Google Sheets"""
    try:
        sheet = conectar_google_sheets()
        if sheet:
            # Actualizar celda A2 con el nuevo valor
            sheet.update_acell('A2', str(nuevo_valor))
            print(f"Consecutivo actualizado en Google Sheets: {nuevo_valor}")
            return True
        # Si falla Google Sheets, usar archivo local como fallback
        with open("consecutivo.txt", "w") as f:
            f.write(str(nuevo_valor))
            print(f"Consecutivo guardado localmente: {nuevo_valor}")
        return True
    except Exception as e:
        print(f"Error guardando consecutivo: {e}")
        return False

def crear_nueva_acr():
    st.markdown("##  Crear Nueva Análisis de Causa Raíz")
    st.markdown("Complete el siguiente formulario para documentar y analizar el problema identificado.")

    # Leer consecutivo automático
    if 'consecutivo' not in st.session_state:
        consecutivo_num = leer_consecutivo()
        st.session_state['consecutivo'] = consecutivo_formateado(consecutivo_num)

    # PRIMERA SECCIÓN: INFORMACIÓN GENERAL
    st.markdown("### 📋 INFORMACIÓN GENERAL")
    st.markdown("---")

    col1, col2, col3 = st.columns(3)

    with col1:
        consecutivo = st.text_input(
            "Consecutivo",
            placeholder="Ej: ACR-001",
            help="Número consecutivo del ACR",
            key="consecutivo"
        )

        fuente_origen = st.text_input(
            "Fuente en la que se origina",
            placeholder="Fuente del problema...",
            key="fuente_origen"
        )

        proceso = st.selectbox(
            "Proceso",
            options=[
                "Direccionamiento Estratégico",
                "Gestión Comercial y de Mercadeo", 
                "Administración de Nómina",
                "Administración de Personal",
                "Selección de Personal",
                "Gestión de Servicio al Cliente",
                "Gestión Administrativa y Financiera",
                "Gestión de Talento Humano",
                "Employer of Record",
                "Gestión Integral",
                "BPO",
                "Outsourcing de tesorería"
            ],
            key="proceso"
        )
    
    with col2:
        cliente = st.text_input(
            "Cliente",
            placeholder="Nombre del cliente...",
            key="cliente"
        )
        
        fecha_incidente = st.date_input(
            "Fecha del incidente",
            value=None,
            help="Fecha en que ocurrió el incidente",
            key="fecha_incidente"
        )
        
        fecha_registro = st.date_input(
            "Fecha de registro",
            value=None,
            help="Fecha de registro del ACR",
            key="fecha_registro"
        )
    
    with col3:
        tipo_accion = st.selectbox(
            "Tipo de acción",
            options=["Correctiva", "De mejora"],
            key="tipo_accion"
        )
        
        tratamiento = st.selectbox(
            "Tratamiento *Solo aplica para Salidas No conformes*",
            options=[
                "No Aplica",
                "Concesión: Autorización para utilizar o liberar una salida que No es conforme con los requisitos especificados",
                "Liberación: Autorización para proseguir con la siguiente etapa de un proceso",
                "Corrección: Acción tomada para eliminar una No Conformidad detectada",
                "Anulación: Acción tomada para declarar inválido la emisión de un documento, factura o similar",
                "Otros"
            ],
            key="tratamiento"
        )
        
        evaluacion_riesgo = st.selectbox(
            "Evaluación del riesgo",
            options=[
                "Riesgo leve - no afecto al cliente - no afecta el contrato (Es poco factible que ocurra)",
                "Riesgo Moderado - insatisfacción del cliente - no afecta el contrato",
                "Riesgo intolerable - afecto la continuidad del contrato",
                "No Aplica"
            ],
            key="evaluacion_riesgo"
        )
    
    descripcion_situacion = st.text_area(
        "Descripción de la situación (actual, potencial o de mejora: Qué, cuándo, dónde, incumplimiento)",
        height=120,
        placeholder="Describa detalladamente la situación...",
        key="descripcion_situacion"
    )
    
    # SEGUNDA SECCIÓN: CORRECCIÓN
    st.markdown("### 🔧 CORRECCIÓN (No aplica para riesgos)")
    st.markdown("---")
    
    st.markdown("**Actividades de corrección inmediata:**")
    
    # Inicializar session state para actividades de corrección
    if 'num_actividades_corr' not in st.session_state:
        st.session_state.num_actividades_corr = 3
    
    # Crear tabla visual para corrección (dinámico, máximo 15 actividades: filas 12-26)
    for i in range(st.session_state.num_actividades_corr):
        st.markdown(f"**Actividad {i+1}:**")
        col_act1, col_act2, col_act3, col_act4, col_act5, col_act6, col_act7 = st.columns([2, 1, 1, 1, 1, 1, 1])
        
        with col_act1:
            st.text_input(f"Actividad", key=f"corr_actividad_{i}", placeholder="Descripción de la actividad...")
        
        with col_act2:
            st.text_input(f"Recursos", key=f"corr_recursos_{i}", placeholder="Recursos necesarios...")
        
        with col_act3:
            st.text_input(f"Responsable", key=f"corr_responsable_{i}", placeholder="Nombre del responsable...")
        
        with col_act4:
            st.number_input(f"Tiempo (Horas)", min_value=0.0, step=0.5, key=f"corr_tiempo_{i}")
        
        with col_act5:
            st.date_input(f"Fecha Inicio", value=None, key=f"corr_fecha_inicio_{i}")
        
        with col_act6:
            st.date_input(f"Fecha Fin", value=None, key=f"corr_fecha_fin_{i}")
        
        with col_act7:
            st.text_input(f"Costo", key=f"corr_costo_{i}", placeholder="Ej: 40000")
        
        if i < st.session_state.num_actividades_corr - 1:
            st.markdown("---")
    
    # Botones para agregar/quitar actividades de corrección
    st.markdown("---")
    col_corr_btn1, col_corr_btn2 = st.columns(2)
    with col_corr_btn1:
        if st.button("➕ Agregar Actividad de Corrección", key="add_actividad_corr"):
            if st.session_state.num_actividades_corr < 15:  # Máximo 15 actividades
                st.session_state.num_actividades_corr += 1
                st.rerun()
            else:
                st.warning("⚠️ Máximo 15 actividades de corrección permitidas")
    with col_corr_btn2:
        if st.button("➖ Quitar Actividad de Corrección", key="del_actividad_corr") and st.session_state.num_actividades_corr > 1:
            st.session_state.num_actividades_corr -= 1
            st.rerun()
    
    # TERCERA SECCIÓN: IDENTIFICACIÓN DE CAUSAS PRINCIPALES
    st.markdown("### 🔍 IDENTIFICACIÓN DE CAUSAS PRINCIPALES")
    st.markdown("---")
    
    # Botón de IA al inicio de esta sección
    col_ia1, col_ia2 = st.columns([1, 2])
    
    with col_ia1:
        if st.button("🚀 Generar con IA", use_container_width=True, type="secondary", key="btn_ia_causas"):
            if descripcion_situacion and descripcion_situacion.strip():
                st.session_state.generando_analisis = True
            else:
                st.warning("⚠️ Primero ingrese la descripción de la situación")
    
    with col_ia2:
        st.markdown("**🤖 Asistente IA:** Genera automáticamente el análisis de causas")
    
    # Procesar generación de análisis si está pendiente
    if st.session_state.get('generando_analisis', False):
        with st.spinner("🔄 Analizando las causas con IA... (esto puede tomar hasta 60 segundos)"):
            analisis_resultado = generar_analisis_ia_simple(descripcion_situacion)
            
            if analisis_resultado:
                st.session_state.analisis_ia_resultado = analisis_resultado
                st.success("✅ ¡Análisis generado exitosamente!")
                st.balloons()
                st.session_state.generando_analisis = False
                st.rerun()
            else:
                st.error("❌ Error al generar análisis. Verifique su conexión a internet e intente nuevamente.")
                st.session_state.generando_analisis = False
    
    # Campo de análisis de causa
    valor_inicial = st.session_state.get('analisis_ia_resultado', '')
    
    if valor_inicial:
        st.info(f"📋 Análisis disponible: {len(valor_inicial)} caracteres")
        with st.expander("👀 Previsualizar análisis generado", expanded=True):
            st.markdown("**Contenido del análisis:**")
            st.text_area("", value=valor_inicial, height=150, disabled=True, key="preview_readonly")
        
        col_copy1, col_copy2 = st.columns(2)
        with col_copy1:
            if st.button("✅ Usar este análisis", type="primary"):
                st.session_state.texto_analisis_causa = valor_inicial
                del st.session_state.analisis_ia_resultado
                st.success("📋 Análisis copiado al formulario")
                st.rerun()
        
        with col_copy2:
            if st.button("🗑️ Descartar análisis"):
                del st.session_state.analisis_ia_resultado
                st.rerun()
    
    analisis_causa = st.text_area(
        "Análisis de causa",
        height=200,
        placeholder="Ingrese el análisis de las causas o genérelo con IA arriba...",
        help="Puede ingresar el análisis manualmente o generarlo automáticamente con IA",
        key="texto_analisis_causa",
        max_chars=5000
    )
    
    if not analisis_causa and valor_inicial:
        st.warning("⚠️ El análisis generado no se pudo cargar en el campo de texto. Usando el análisis mostrado arriba.")
        analisis_causa = valor_inicial
    
    # Inicializar session state para causas inmediatas y raíz
    if 'num_causas_inmediatas' not in st.session_state:
        st.session_state.num_causas_inmediatas = 3
    if 'num_causas_raiz' not in st.session_state:
        st.session_state.num_causas_raiz = 3
    
    # Causas Inmediatas, básicas y/o gerenciales (dinámico, máximo 5)
    st.markdown("**Causa(s) Inmediata, básica y/o gerenciales:**")
    
    cols_inmediatas = st.columns(st.session_state.num_causas_inmediatas)
    for i in range(st.session_state.num_causas_inmediatas):
        with cols_inmediatas[i]:
            st.text_area(f"Causa {i+1}", height=100, key=f"causa_inmediata_{i+1}")
    
    # Botones para agregar/quitar causas inmediatas
    col_inm_btn1, col_inm_btn2 = st.columns(2)
    with col_inm_btn1:
        if st.button("➕ Agregar Causa Inmediata", key="add_causa_inmediata"):
            if st.session_state.num_causas_inmediatas < 5:
                st.session_state.num_causas_inmediatas += 1
                st.rerun()
            else:
                st.warning("⚠️ Máximo 5 causas inmediatas permitidas")
    with col_inm_btn2:
        if st.button("➖ Quitar Causa Inmediata", key="del_causa_inmediata") and st.session_state.num_causas_inmediatas > 1:
            st.session_state.num_causas_inmediatas -= 1
            st.rerun()
    
    st.markdown("---")
    
    # Causas Raíz (dinámico, máximo 5)
    st.markdown("**Causa(s) Raíz(s):**")
    
    cols_raiz = st.columns(st.session_state.num_causas_raiz)
    for i in range(st.session_state.num_causas_raiz):
        with cols_raiz[i]:
            st.text_area(f"Causa Raíz {i+1}", height=100, key=f"causa_raiz_{i+1}")
    
    # Botones para agregar/quitar causas raíz
    col_raiz_btn1, col_raiz_btn2 = st.columns(2)
    with col_raiz_btn1:
        if st.button("➕ Agregar Causa Raíz", key="add_causa_raiz"):
            if st.session_state.num_causas_raiz < 5:
                st.session_state.num_causas_raiz += 1
                st.rerun()
            else:
                st.warning("⚠️ Máximo 5 causas raíz permitidas")
    with col_raiz_btn2:
        if st.button("➖ Quitar Causa Raíz", key="del_causa_raiz") and st.session_state.num_causas_raiz > 1:
            st.session_state.num_causas_raiz -= 1
            st.rerun()
    
    # CUARTA SECCIÓN: PLAN DE ACCIÓN
    st.markdown("### 💡 PLAN DE ACCIÓN")
    st.markdown("---")
    
    st.markdown("**Nuevo modelo: Tabla de actividades con causas asociadas**")
    
    # 1. CONFIGURACIÓN DE CAUSAS
    st.markdown("#### 📋 **Paso 1: Definir Causas**")
    
    # Inicializar session state para causas
    if 'num_causas_pa' not in st.session_state:
        st.session_state.num_causas_pa = 3
    
    # Botones para agregar/quitar causas
    col_causas_btn1, col_causas_btn2 = st.columns(2)
    with col_causas_btn1:
        if st.button("✚ Agregar Causa", key="add_causa_pa"):
            st.session_state.num_causas_pa += 1
            st.rerun()
    with col_causas_btn2:
        if st.button("━ Quitar Causa", key="del_causa_pa") and st.session_state.num_causas_pa > 1:
            st.session_state.num_causas_pa -= 1
            st.rerun()
    
    # Campos para las causas
    causas_disponibles = []
    for i in range(st.session_state.num_causas_pa):
        causa_text = st.text_area(
            f"**Causa Asociada {i+1}:**", 
            height=80,
            key=f"pa_causa_def_{i}",
            placeholder=f"Describe la causa asociada {i+1}..."
        )
        if causa_text:
            causas_disponibles.append(f"Causa {i+1}")
    
    st.markdown("---")
    
    # 2. TABLA DE ACTIVIDADES
    st.markdown("#### 🎯 **Paso 2: Definir Actividades**")
    
    # Inicializar session state para actividades
    if 'num_actividades_pa' not in st.session_state:
        st.session_state.num_actividades_pa = 1
    
    # Tabla de actividades
    for i in range(st.session_state.num_actividades_pa):
        st.markdown(f"**ACTIVIDAD {i+1}:**")
        
        # Fila 1: Actividad y Causas Asociadas
        col_act1, col_act2 = st.columns([2, 1])
        
        with col_act1:
            actividad = st.text_area(
                "Descripción de la actividad",
                height=80,
                key=f"pa_actividad_nueva_{i}",
                placeholder="Describe la actividad a desarrollar..."
            )
        
        with col_act2:
            if causas_disponibles:
                causas_asociadas = st.multiselect(
                    "Causas asociadas a esta actividad",
                    options=causas_disponibles,
                    key=f"pa_causas_asociadas_{i}",
                    help="Selecciona una o varias causas que esta actividad ayudará a resolver"
                )
            else:
                st.warning("⚠️ Primero define las causas en el Paso 1")
                causas_asociadas = []
        
        # Fila 2: Campos de gestión
        col_gest1, col_gest2, col_gest3 = st.columns(3)
        
        with col_gest1:
            responsable_ej = st.text_input(
                "Responsable Ejecución",
                key=f"pa_resp_ej_nueva_{i}",
                placeholder="Responsable de ejecutar..."
            )
            tiempo = st.number_input(
                "Tiempo (Horas)",
                min_value=0.0,
                step=0.5,
                key=f"pa_tiempo_nueva_{i}"
            )
            
        with col_gest2:
            costo = st.text_input(
                "Costo",
                key=f"pa_costo_nueva_{i}",
                placeholder="Ej: 40000"
            )
            fecha_inicio = st.date_input(
                "Fecha Inicio",
                value=None,
                key=f"pa_fecha_inicio_nueva_{i}"
            )
            
        with col_gest3:
            fecha_fin = st.date_input(
                "Fecha Fin",
                value=None,
                key=f"pa_fecha_fin_nueva_{i}"
            )
            responsable_seg = st.text_input(
                "Responsable Seguimiento",
                key=f"pa_resp_seg_nueva_{i}",
                placeholder="Responsable del seguimiento..."
            )
        
        # Fila 3: Campos de seguimiento
        col_seg1, col_seg2, col_seg3 = st.columns(3)
        
        with col_seg1:
            fecha_seguimiento = st.date_input(
                "Fecha Seguimiento",
                value=None,
                key=f"pa_fecha_seg_nueva_{i}"
            )
            
        with col_seg2:
            estado = st.selectbox(
                "Estado",
                options=["", "Abierta", "Cerrada", "Parcial"],
                key=f"pa_estado_nueva_{i}"
            )
            
        with col_seg3:
            costo_seguimiento = st.text_input(
                "Costo Seguimiento",
                key=f"pa_costo_seg_nueva_{i}",
                placeholder="Ej: 40000"
            )
        
        # Evidencia (campo amplio)
        evidencia = st.text_input(
            "Evidencia de Verificación",
            key=f"pa_evidencia_nueva_{i}",
            placeholder="Evidencia que demuestre el cumplimiento..."
        )
        
        # Mostrar causas asociadas si hay
        if causas_asociadas:
            st.info(f"📌 Esta actividad está asociada a: {', '.join(causas_asociadas)}")
        
        if i < st.session_state.num_actividades_pa - 1:
            st.markdown("---")

    # Botones para agregar/quitar actividades debajo de la última actividad
    st.markdown(":heavy_minus_sign:" * 30)  # Separador visual
    col_act_btn1, col_act_btn2 = st.columns(2)
    with col_act_btn1:
        if st.button("✚ Agregar Actividad", key="add_actividad_pa"):
            st.session_state.num_actividades_pa += 1
            st.rerun()
    with col_act_btn2:
        if st.button("━ Quitar Actividad", key="del_actividad_pa") and st.session_state.num_actividades_pa > 1:
            st.session_state.num_actividades_pa -= 1
            st.rerun()
    
    # QUINTA SECCIÓN: COSTOS ASOCIADOS A LA ACR
    st.markdown("### 💰 COSTOS ASOCIADOS A LA ACR")
    st.markdown("---")
    
    col_costo1, col_costo2 = st.columns(2)
    
    with col_costo1:
        costo_correccion = st.text_input(
            "Costos de la corrección (Reproceso Interno)",
            key="costo_correccion",
            placeholder="Ej: 40000"
        )
        
        costo_reputacional = st.text_input(
            "Costo daño reputacional (5-10% del valor del contrato con el cliente afectado)",
            key="costo_reputacional",
            placeholder="Ej: 40000"
        )
        
        costo_acciones = st.text_input(
            "Costos de las acciones correctivas",
            key="costo_acciones",
            placeholder="Ej: 40000"
        )
        
        multas_sanciones = st.text_input(
            "Multas / Sanciones",
            key="multas_sanciones",
            placeholder="Ej: 40000"
        )
    
    with col_costo2:
        costo_seguimiento = st.text_input(
            "Costos de seguimiento",
            key="costo_seguimiento",
            placeholder="Ej: 40000"
        )
        
        otros_costos_internos = st.text_input(
            "Otros costos internos asociados a la NC, no cuantificados",
            key="otros_costos_internos",
            placeholder="Ej: 40000"
        )
        
        descuentos_cliente = st.text_input(
            "Descuentos realizados al cliente",
            key="descuentos_cliente",
            placeholder="Ej: 40000"
        )
        
        otros_costos = st.text_input(
            "Otros costos",
            key="otros_costos",
            placeholder="Ej: 40000"
        )
    
    # Botones de acción
    st.markdown("---")
    col_btn1, col_btn2 = st.columns([2, 2])
    
    with col_btn1:
        if st.button("📊 Generar Reporte Excel", use_container_width=True, type="primary"):
            # Validar todos los campos de información general
            campos_info = {
                'Consecutivo': consecutivo,
                'Fuente en la que se origina': fuente_origen,
                'Proceso': proceso,
                'Cliente': st.session_state.get('cliente', ''),
                'Fecha del incidente': st.session_state.get('fecha_incidente', ''),
                'Fecha de registro': st.session_state.get('fecha_registro', ''),
                'Tipo de acción': tipo_accion,
                'Tratamiento': tratamiento,
                'Evaluación del riesgo': evaluacion_riesgo,
                'Descripción de la situación': descripcion_situacion
            }
            campos_faltantes = [nombre for nombre, valor in campos_info.items() if not valor or (isinstance(valor, str) and not valor.strip())]
            if campos_faltantes:
                st.error("❌ Por favor, complete los siguientes campos obligatorios:")
                for campo in campos_faltantes:
                    st.warning(f"- {campo}")
            else:
                with st.spinner("Generando reporte..."):
                    excel_data = generar_excel_acr_completo()
                    if excel_data:
                        st.session_state['excel_generado'] = excel_data
                        st.session_state['consecutivo_generado'] = consecutivo
                        st.success("✅ ¡Reporte generado exitosamente!")
                        # Incrementar y guardar consecutivo solo si la generación fue exitosa (sin recargar la app)
                        try:
                            consecutivo_actual = int(st.session_state['consecutivo'])
                            nuevo_consecutivo = consecutivo_actual + 1
                            guardar_consecutivo(nuevo_consecutivo)
                        except Exception as e:
                            st.warning(f"No se pudo actualizar el consecutivo automáticamente: {e}")

        # Mostrar área de nombre y botón de descarga si ya hay un excel generado
        if 'excel_generado' in st.session_state and 'consecutivo_generado' in st.session_state:
            if 'nombre_archivo_usuario' not in st.session_state:
                st.session_state['nombre_archivo_usuario'] = ""
            nombre_usuario = st.text_area(
                "Nombre personalizado para el archivo (opcional)",
                value=st.session_state['nombre_archivo_usuario'],
                placeholder="Ejemplo: CLIENTE XYZ, INCIDENTE 2025, etc.",
                key="nombre_archivo_usuario"
            )
            nombre_final = f"{st.session_state['consecutivo_generado']} ACCIONES CORRECTIVAS Y DE MEJORA"
            if st.session_state['nombre_archivo_usuario'].strip():
                nombre_final += f" {st.session_state['nombre_archivo_usuario'].strip()}"
            nombre_final += ".xlsx"
            st.info(f"El archivo se descargará como: \n**{nombre_final}**")
            st.download_button(
                label="⬇️ Descargar Reporte ACR",
                data=st.session_state['excel_generado'],
                file_name=nombre_final,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            # Botón para enviar por correo al encargado
            # st.caption("⚠️ Límites: 1 correo cada 5 min | Máximo 10 correos/día")
            # if st.button("📧 Enviar ACR al encargado"):
            #     exito = enviar_acr_smtp_env(
            #         st.session_state['excel_generado'],
            #         nombre_final
            #     )
            #     if exito:
            #         st.success("✅ Correo enviado correctamente al encargado.")
    
    with col_btn2:
        if st.button("🔄 Limpiar Formulario", use_container_width=True):
            # Limpiar todos los campos del session_state
            keys_to_clear = [key for key in st.session_state.keys() if not key.startswith('_')]
            for key in keys_to_clear:
                del st.session_state[key]
            st.rerun()

    # def cargar_acr_existente():
    #     st.markdown("## 📤 Consolidar ACR en Excel Maestro")
    #     st.markdown("Transfiera los datos de un ACR individual al registro histórico maestro.")
    #     
    #     # Información del proceso
    #     with st.expander("ℹ️ ¿Cómo funciona este proceso?", expanded=True):
    #         st.markdown("""
    #         **Pasos del proceso:**
    #         1. Seleccione el archivo Excel del ACR individual
    #         2. Seleccione el archivo Excel maestro (historial)
    #         3. El sistema consolidará automáticamente los datos
    #         4. Descargue el Excel maestro actualizado
    #         
    #         **Nota:** Los datos se agregarán al final del historial existente.
    #         """)
    #     
    #     st.markdown("---")
    #     
    #     col1, col2 = st.columns(2)
    #     
    #     with col1:
    #         st.markdown("### 📄 Archivo ACR Individual")
    #         archivo_acr = st.file_uploader(
    #             "Seleccione el archivo Excel del ACR",
    #             type=['xlsx', 'xls'],
    #             help="Archivo generado por el sistema de creación de ACR",
    #             key="acr_file"
    #         )
    #         
    #         if archivo_acr:
    #             st.success(f"✓ Archivo cargado: {archivo_acr.name}")
    #     
    #     with col2:
    #         st.markdown("### 📚 Excel Maestro (Historial)")
    #         archivo_maestro = st.file_uploader(
    #             "Seleccione el Excel maestro",
    #             type=['xlsx', 'xls'],
    #             help="Archivo que contiene el historial de todos los ACR",
    #             key="maestro_file"
    #         )
    #         
    #         if archivo_maestro:
    #             st.success(f"✓ Archivo cargado: {archivo_maestro.name}")
    #     
    #     st.markdown("---")
    #     
    #     if archivo_acr and archivo_maestro:
    #         col_preview1, col_preview2 = st.columns(2)
    #         
    #         with col_preview1:
    #             with st.expander("👁️ Vista previa del ACR"):
    #                 try:
    #                     df_preview = pd.read_excel(archivo_acr)
    #                     st.dataframe(df_preview, use_container_width=True)
    #                 except Exception as e:
    #                     st.error(f"Error al leer el archivo: {str(e)}")
    #         
    #         with col_preview2:
    #             with st.expander("👁️ Vista previa del Maestro"):
    #                 try:
    #                     df_maestro_preview = pd.read_excel(archivo_maestro)
    #                     st.dataframe(df_maestro_preview.tail(5), use_container_width=True)
    #                     st.info(f"Total de registros: {len(df_maestro_preview)}")
    #                 except Exception as e:
    #                     st.error(f"Error al leer el archivo: {str(e)}")
    #         
    #         st.markdown("---")
    #         
    #         col_btn = st.columns([1, 2, 1])[1]
    #         with col_btn:
    #             if st.button("🔄 Consolidar Datos", use_container_width=True, type="primary"):
    #                 with st.spinner("Procesando archivos..."):
    #                     resultado = procesar_archivos_excel(archivo_acr, archivo_maestro)
    #                     if resultado:
    #                         st.success("✅ ¡Datos consolidados exitosamente!")
    #                         st.download_button(
    #                             label="⬇️ Descargar Excel Maestro Actualizado",
    #                             data=resultado,
    #                             file_name=f"Maestro_ACR_SolutionsPayroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    #                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #                             use_container_width=True
    #                         )

def mostrar_informacion_sistema():
    st.markdown("## 📊 Información del Sistema")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        logo_path = "syp logo.png"
        if os.path.exists(logo_path):
            with open(logo_path, "rb") as image_file:
                encoded = base64.b64encode(image_file.read()).decode()
            img_html = f"<img src='data:image/png;base64,{encoded}' width='32' style='vertical-align:middle;margin-right:8px;'/>"
        else:
            img_html = ""
        st.markdown(f"""
        <div class="stat-card">
            <h3>{img_html}Empresa</h3>
            <p><strong>Solutions & Payroll</strong></p>
            <p>Sistema de Gestión ACR</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="stat-card">
            <h3>📦 Versión</h3>
            <p><strong>2.0</strong></p>
            <p>Última actualización: Oct 2025</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="stat-card">
            <h3>🤖 IA Integrada</h3>
            <p><strong>Google Gemini</strong></p>
            <p>Análisis automático de causas</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    st.markdown("### 📖 Manual de Usuario")
    
    with st.expander("1️⃣ Crear Nueva ACR", expanded=True):
        st.markdown("""
        **Proceso para crear un nuevo análisis:**
        - Complete la descripción detallada del problema
        - Especifique el área afectada y el responsable
        - Utilice la IA para generar el análisis de causa raíz automáticamente
        - Defina las acciones de mejora y la fecha compromiso
        - Genere y descargue el reporte en Excel
        """)
    
    with st.expander("2️⃣ Análisis con IA"):
        st.markdown("""
        **Características del asistente de IA:**
        - Genera análisis de los 5 porqués automáticamente
        - Identifica la causa raíz del problema
        - Proporciona un análisis estructurado y profesional
        - Requiere configuración de API key de Gemini
        """)

def generar_analisis_ia_simple(descripcion_problema):
    """Versión simplificada con debug para generar análisis"""
    try:
        print(f"DEBUG: Iniciando análisis IA con descripción: {descripcion_problema[:50]}...")
        
        # Cargar API key
        load_dotenv()
        api_key = os.getenv("GEMINI_API_KEY")
        print(f"DEBUG: API Key encontrada: {bool(api_key)} (longitud: {len(api_key) if api_key else 0})")
        
        if not api_key:
            print("DEBUG: No se encontró API key en environment, intentando leer .env...")
            try:
                with open('.env', 'r') as f:
                    content = f.read()
                    for line in content.split('\n'):
                        if line.startswith('GEMINI_API_KEY='):
                            api_key = line.split('=', 1)[1].strip()
                            print(f"DEBUG: API Key leída de .env: {bool(api_key)} (longitud: {len(api_key) if api_key else 0})")
                            break
            except Exception as e:
                print(f"DEBUG: Error leyendo .env: {e}")
                return None
        
        if not api_key:
            print("DEBUG: No se pudo obtener API key")
            return None
        
        # Prompt específico para análisis de causas (optimizado para respuestas concretas con ejemplos)
        prompt = f"""Analiza la siguiente situación y proporciona un análisis de causa raíz específico y directo, siguiendo el formato de los ejemplos proporcionados.

**EJEMPLO 1:**
SITUACIÓN: "El cliente DISTRITECH COLOMBIA SAS decidió finalizar el contrato suscrito con S&P debido a reiterados errores ocurridos durante la prestación de los servicios de Administración de Personal y Administración de Nómina en los años 2024 y 2025, los cuales afectaron la calidad, oportunidad y confiabilidad de la información entregada. Entre las inconsistencias evidenciadas se encuentran: errores en la retención en la fuente a tres empleados durante el año 2024; liquidación incorrecta de la planilla de cesantías de 2024 para un empleado; reporte erróneo de la información exógena y de los certificados de ingresos y retenciones de 2024; fallas en la configuración del usuario de autoconsulta para una empleada en dos ocasiones; afiliaciones de empleados a un empleador equivocado; y una afiliación incorrecta a la ARL en marzo de 2025, pese a que el cliente había realizado el cambio de aseguradora en enero del mismo año."

ANÁLISIS DE LOS 5 PORQUÉS:
¿Por qué 1? ¿Por qué se presentaron diferentes errores que derivaron en la finalización del contrato por parte del cliente? - No tenemos como garantizar que las políticas de doble verificación establecidas en los procedimientos de la compañía se cumplan
¿Por qué 2? ¿Por qué no se garantiza el cumplimiento de las políticas de doble verificación? - No contamos con puntos de control críticos definidos dentro de los procedimientos asociados a ADP y ADN
¿Por qué 3? ¿Por qué no hay puntos de control críticos definidos? - Porque no existen mecanismos de trazabilidad claros que evidencien la verificación realizada en los puntos críticos de cada proceso, identificando posibles errores humanos
¿Por qué 4? ¿Por qué no existen mecanismos de trazabilidad claros? - Porque aunque se identificaron errores humanos recurrentes no se aplicó el proceso disciplinario correspondiente en su debido momento
¿Por qué 5? ¿Por qué no se aplicó el proceso disciplinario en su debido momento? - Porque no hay una cultura organizacional que integre el cumplimiento de procedimiento con la aplicación de correctivos disciplinarios, lo que permitió la reincidencia de errores y la pérdida de confianza del cliente

CAUSAS INMEDIATAS:
- No tenemos como garantizar que las políticas de doble verificación establecidas en los procedimientos de la compañía se cumplan
- No contamos con puntos de control críticos definidos dentro de los procedimientos asociados a ADP y ADN
- No existen mecanismos de trazabilidad claros que evidencien la verificación realizada en los puntos críticos de cada proceso, identificando posibles errores humanos

CAUSAS RAÍZ:
- Aunque se identificaron errores humanos recurrentes no se aplicó el proceso disciplinario correspondiente en su debido momento para evitar su repetición
- No hay una cultura organizacional que integre el cumplimiento de procedimiento con la aplicación de correctivos disciplinarios, lo que permitió la reincidencia de errores y la pérdida de confianza del cliente

---

**EJEMPLO 2:**
SITUACIÓN: "Se evidenció que el impuesto de Industria y Comercio (ICA) correspondiente no fue pagado oportunamente. Aunque el 13 de junio se realizó la solicitud de pago a través de WhatsApp, no se adjuntó el recibo oficial del impuesto, lo cual impidió que el equipo de tesorería ejecutara el pago. El 20 de octubre, el contador identificó que el impuesto seguía pendiente, generándose intereses por mora. No se efectuó seguimiento por parte del outsourcing de tesorería ni del cliente, lo que ocasionó el incumplimiento del pago en los tiempos establecidos."

ANÁLISIS DE LOS 5 PORQUÉS:
¿Por qué 1? ¿Por qué no se realizó el pago del impuesto? - Porque no se recibió el recibo de pago junto con la solicitud que se realizó por WhatsApp, y adicionalmente, no se realizó seguimiento por ninguna de las dos partes
¿Por qué 2? ¿Por qué no se adjuntó el recibo de pago ni se realizó seguimiento? - Porque no existía un mecanismo claro de control o verificación dentro del proceso que consolidara los pagos que se deben realizar por PSE como el pago de impuestos
¿Por qué 3? ¿Por qué no existía un mecanismo de control o registro para los pagos por PSE? - Porque el formato o archivo utilizado para reportar y hacer seguimiento a los pagos no contemplaba inicialmente los pagos de impuestos realizados por este medio
¿Por qué 4? ¿Por qué el formato no contemplaba los pagos por PSE? - Porque no se había unificado la gestión de todos los tipos de pagos (bancarios y por PSE) dentro de un mismo registro o procedimiento que facilitara el control conjunto

CAUSAS INMEDIATAS:
- No se recibió el recibo de pago junto con la solicitud que se realizó por WhatsApp, y adicionalmente, no se realizó seguimiento por ninguna de las dos partes
- No existía un mecanismo claro de control o verificación dentro del proceso que consolidara los pagos que se deben realizar por PSE como el pago de impuestos

CAUSAS RAÍZ:
- El formato o archivo utilizado para reportar y hacer seguimiento a los pagos no contemplaba inicialmente los pagos de impuestos realizados por este medio
- No se había unificado la gestión de todos los tipos de pagos (bancarios y por PSE) dentro de un mismo registro o procedimiento que facilitara el control conjunto

---

**AHORA ANALIZA ESTE CASO:**
SITUACIÓN: {descripcion_problema}

Proporciona ÚNICAMENTE:

1. ANÁLISIS DE LOS 5 PORQUÉS:
¿Por qué 1? 
¿Por qué 2? 
¿Por qué 3? 
¿Por qué 4? 
¿Por qué 5? 

2. CAUSAS INMEDIATAS (2-3 causas específicas y concretas):
- 
- 
- 

3. CAUSAS RAÍZ (2-3 causas fundamentales y específicas):
- 
- 

INSTRUCCIONES:
- Sigue el mismo formato y nivel de detalle de los ejemplos anteriores
- Sé específico y directo, enfocándote en problemas sistémicos y procedimentales
- Identifica fallas en procesos, controles, seguimiento y cultura organizacional
- NO incluyas explicaciones largas, ejemplos adicionales ni recomendaciones
- Las causas deben ser claras, concretas y orientadas a la mejora de procesos"""
        
        # Llamada API
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        print(f"DEBUG: URL API: {url[:80]}...")
        
        headers = {"Content-Type": "application/json"}
        data = {
            "contents": [{
                "parts": [{
                    "text": prompt
                }]
            }]
        }
        
        print(f"DEBUG: Enviando request a Gemini...")
        response = requests.post(url, headers=headers, json=data, timeout=60)
        print(f"DEBUG: Respuesta recibida - Status: {response.status_code}")
        
        if response.status_code == 200:
            result = response.json()
            print(f"DEBUG: Respuesta JSON recibida correctamente")
            if 'candidates' in result and len(result['candidates']) > 0:
                texto = result['candidates'][0]['content']['parts'][0]['text']
                # Limpiar el texto de caracteres problemáticos
                texto_limpio = texto.strip()
                # Asegurar que no hay caracteres de control problemáticos
                texto_limpio = ''.join(char for char in texto_limpio if ord(char) >= 32 or char in '\n\r\t')
                print(f"DEBUG: Análisis generado exitosamente ({len(texto_limpio)} caracteres)")
                return texto_limpio
            else:
                print(f"DEBUG: No se encontraron candidates en la respuesta: {result}")
        else:
            print(f"DEBUG: Error en API - Status: {response.status_code}, Response: {response.text}")
        
        return None
        
    except Exception as e:
        print(f"DEBUG: Excepción en generar_analisis_ia_simple: {type(e).__name__}: {e}")
        return None

def generar_analisis_ia(descripcion_problema):
    """Genera el análisis de los 5 porqués usando la API de Gemini"""
    try:
        # Mostrar estado inicial
        st.write("🔍 Iniciando análisis IA...")
        
        # Recargar variables de entorno
        load_dotenv()
        
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key or len(api_key) != 39:
            # Leer directamente del archivo .env
            try:
                with open('.env', 'r') as f:
                    content = f.read()
                    for line in content.split('\n'):
                        if line.startswith('GEMINI_API_KEY='):
                            api_key = line.split('=', 1)[1].strip()
                            break
            except Exception as e:
                st.error(f"❌ Error leyendo archivo .env: {e}")
                return None
        
        if not api_key or len(api_key) != 39:
            st.error("❌ API Key no válida")
            return None
        
        st.write("✅ API Key configurada")
        
        # Prompt simplificado
        prompt = f"""Realiza un análisis de los 5 porqués para este problema:

Problema: {descripcion_problema}

Estructura tu respuesta así:

1. ¿Por qué ocurrió este problema?
   Respuesta: [tu análisis]

2. ¿Por qué [causa del punto 1]?
   Respuesta: [tu análisis]

3. ¿Por qué [causa del punto 2]?
   Respuesta: [tu análisis]

4. ¿Por qué [causa del punto 3]?
   Respuesta: [tu análisis]

5. ¿Por qué [causa del punto 4]?
   Respuesta: [tu análisis]

CAUSA RAÍZ: [La causa fundamental]"""
        
        st.write("🚀 Enviando petición a Gemini...")
        
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        
        headers = {"Content-Type": "application/json"}
        data = {
            "contents": [{
                "parts": [{
                    "text": prompt
                }]
            }]
        }
        
        # Petición con timeout
        response = requests.post(url, headers=headers, json=data, timeout=60)
        
        st.write(f"📡 Respuesta HTTP: {response.status_code}")
        
        if response.status_code == 200:
            result = response.json()
            st.write("✅ JSON parseado correctamente")
            
            if 'candidates' in result and len(result['candidates']) > 0:
                texto_respuesta = result['candidates'][0]['content']['parts'][0]['text']
                st.write(f"📝 Texto recibido: {len(texto_respuesta)} caracteres")
                st.write(f"🔍 Primeros 100 caracteres: {texto_respuesta[:100]}")
                
                # Asegurar que el texto esté en UTF-8
                if isinstance(texto_respuesta, str):
                    return texto_respuesta
                else:
                    return str(texto_respuesta)
            else:
                st.error("❌ Respuesta vacía de la API")
                st.write(f"🔍 Estructura de respuesta: {result}")
                return None
        else:
            st.error(f"❌ Error API: {response.status_code}")
            st.write(f"🔍 Respuesta: {response.text}")
            return None
            
    except requests.exceptions.Timeout:
        st.error("❌ Timeout: La petición tardó más de 60 segundos")
        return None
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        return None

def validar_campos_nuevos(consecutivo, descripcion_situacion):
    """Valida que los campos obligatorios del nuevo formato estén completos"""
    campos_obligatorios = [consecutivo, descripcion_situacion]
    return all(campo and campo.strip() for campo in campos_obligatorios)

def validar_campos(descripcion, area, responsable, acciones, fecha):
    """Valida que todos los campos obligatorios estén completos"""
    campos_obligatorios = [descripcion, area, responsable, acciones]
    return all(campo.strip() for campo in campos_obligatorios) and fecha is not None

def formatear_fecha(fecha):
    """Convierte una fecha al formato 1/10/2025"""
    if fecha is None:
        return ""
    if isinstance(fecha, str):
        return fecha
    # Formato: día/mes/año sin ceros a la izquierda en día y mes
    return f"{fecha.day}/{fecha.month}/{fecha.year}"

def convertir_a_numero(valor):
    """Convierte un valor de texto a número de forma segura"""
    if valor is None or valor == "":
        return 0
    if isinstance(valor, (int, float)):
        return valor
    # Si es string, intentar convertir
    try:
        # Eliminar espacios, comas y puntos de miles
        valor_limpio = str(valor).replace(' ', '').replace(',', '').replace('.', '')
        return float(valor_limpio) if valor_limpio else 0
    except:
        return 0

def generar_excel_acr_completo():
    """Genera el archivo Excel completo con todos los datos del formulario usando las celdas exactas especificadas"""
    try:
        # Cargar el formato base
        workbook = openpyxl.load_workbook("Formato ACR - limpio.xlsx")
        sheet = workbook.active
        
        # Función helper para escribir de forma segura en celdas
        def escribir_celda_segura(celda_ref, valor):
            try:
                cell = sheet[celda_ref]
                if not isinstance(cell, openpyxl.cell.MergedCell):
                    sheet[celda_ref] = valor
            except Exception as e:
                print(f"DEBUG: Error escribiendo en {celda_ref}: {e}")
        
        # PRIMERA SECCIÓN: INFORMACIÓN GENERAL - Mapeo exacto
        consecutivo = st.session_state.get('consecutivo', '')
        fuente_origen = st.session_state.get('fuente_origen', '')
        proceso = st.session_state.get('proceso', '')
        cliente = st.session_state.get('cliente', '')
        fecha_incidente = st.session_state.get('fecha_incidente', '')
        fecha_registro = st.session_state.get('fecha_registro', '')
        tipo_accion = st.session_state.get('tipo_accion', '')
        tratamiento = st.session_state.get('tratamiento', '')
        evaluacion_riesgo = st.session_state.get('evaluacion_riesgo', '')
        descripcion_situacion = st.session_state.get('descripcion_situacion', '')
        
        # Llenar campos primera sección
        if consecutivo:
            escribir_celda_segura('D4', consecutivo)
        if fuente_origen:
            escribir_celda_segura('D5', fuente_origen)
        if proceso:
            escribir_celda_segura('I4', proceso)
        if cliente:
            escribir_celda_segura('I5', cliente)
        if fecha_incidente:
            escribir_celda_segura('O4', formatear_fecha(fecha_incidente))
        if fecha_registro:
            escribir_celda_segura('O5', formatear_fecha(fecha_registro))
        if tipo_accion:
            escribir_celda_segura('V4', tipo_accion)
        if tratamiento:
            escribir_celda_segura('V5', tratamiento)
        if evaluacion_riesgo:
            escribir_celda_segura('Z4', evaluacion_riesgo)
        if descripcion_situacion:
            escribir_celda_segura('A8', descripcion_situacion)
        
        # SEGUNDA SECCIÓN: CORRECCIÓN (15 actividades: A12-A26)
        num_actividades_corr_escritas = st.session_state.get('num_actividades_corr', 3)
        
        for i in range(15):  # 15 actividades (0-14)
            row = 12 + i  # Filas 12-26
            actividad = st.session_state.get(f'corr_actividad_{i}', '')
            recursos = st.session_state.get(f'corr_recursos_{i}', '')
            responsable = st.session_state.get(f'corr_responsable_{i}', '')
            tiempo = st.session_state.get(f'corr_tiempo_{i}', 0)
            fecha_inicio = st.session_state.get(f'corr_fecha_inicio_{i}', '')
            fecha_fin = st.session_state.get(f'corr_fecha_fin_{i}', '')
            costo = convertir_a_numero(st.session_state.get(f'corr_costo_{i}', 0))
            
            # Si la actividad está vacía y está más allá del número visible, ocultar fila
            if not actividad and i >= num_actividades_corr_escritas:
                sheet.row_dimensions[row].hidden = True
            else:
                # Escribir datos si existen
                if actividad:
                    escribir_celda_segura(f'A{row}', actividad)  # ACTIVIDADES A12-A26
                if recursos:
                    escribir_celda_segura(f'J{row}', recursos)   # RECURSOS J12-J26
                if responsable:
                    escribir_celda_segura(f'M{row}', responsable) # RESPONSABLES M12-M26
                if tiempo > 0:
                    escribir_celda_segura(f'O{row}', tiempo)     # TIEMPO O12-O26
                if fecha_inicio:
                    escribir_celda_segura(f'R{row}', formatear_fecha(fecha_inicio)) # FECHA INICIO R12-R26
                if fecha_fin:
                    escribir_celda_segura(f'V{row}', formatear_fecha(fecha_fin))    # FECHA FIN V12-V26
                if costo > 0:
                    escribir_celda_segura(f'Y{row}', costo)      # COSTO Y12-Y26
        
        # TERCERA SECCIÓN: ANÁLISIS DE CAUSA
        analisis_causa = st.session_state.get('texto_analisis_causa', '')
        if analisis_causa:
            escribir_celda_segura('A18', analisis_causa)  # Análisis de causa
        
        # CAUSAS INMEDIATAS (dinámico, hasta 5 causas en filas 30, 32, 34, 36, 38)
        filas_inmediatas = [30, 32, 34, 36, 38]
        for i in range(5):
            causa_inmediata = st.session_state.get(f'causa_inmediata_{i+1}', '')
            if causa_inmediata:
                escribir_celda_segura(f'F{filas_inmediatas[i]}', causa_inmediata)
        
        # CAUSAS RAÍZ (dinámico, hasta 5 causas en filas 40, 42, 44, 46, 48)
        filas_raiz = [40, 42, 44, 46, 48]
        for i in range(5):
            causa_raiz = st.session_state.get(f'causa_raiz_{i+1}', '')
            if causa_raiz:
                escribir_celda_segura(f'F{filas_raiz[i]}', causa_raiz)
        
        # CUARTA SECCIÓN: PLAN DE ACCIÓN (Nuevo modelo: A53-A72)
        
        # Obtener datos del nuevo modelo
        num_causas_pa = st.session_state.get('num_causas_pa', 3)
        num_actividades_pa = st.session_state.get('num_actividades_pa', 1)
        
        # Primero, deshacemos todas las combinaciones existentes en el rango 53-72
        rangos_a_eliminar = []
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.min_row >= 53 and merged_range.max_row <= 72:
                rangos_a_eliminar.append(merged_range)
        
        for rango in rangos_a_eliminar:
            sheet.unmerge_cells(str(rango))
        
        # 1. Crear mapa de causas y sus textos
        causas_map = {}
        for i in range(num_causas_pa):
            causa_text = st.session_state.get(f'pa_causa_def_{i}', '')
            if causa_text:
                causas_map[f"Causa {i+1}"] = causa_text
        
        # 2. Crear mapa de actividades y sus causas asociadas
        actividades_map = {}
        for i in range(num_actividades_pa):
            actividad_text = st.session_state.get(f'pa_actividad_nueva_{i}', '')
            causas_asociadas = st.session_state.get(f'pa_causas_asociadas_{i}', [])
            
            if actividad_text and causas_asociadas:
                actividades_map[i] = {
                    'texto': actividad_text,
                    'causas': causas_asociadas,
                    'responsable_ej': st.session_state.get(f'pa_resp_ej_nueva_{i}', ''),
                    'tiempo': st.session_state.get(f'pa_tiempo_nueva_{i}', 0),
                    'costo': convertir_a_numero(st.session_state.get(f'pa_costo_nueva_{i}', 0)),
                    'fecha_inicio': st.session_state.get(f'pa_fecha_inicio_nueva_{i}', ''),
                    'fecha_fin': st.session_state.get(f'pa_fecha_fin_nueva_{i}', ''),
                    'responsable_seg': st.session_state.get(f'pa_resp_seg_nueva_{i}', ''),
                    'fecha_seguimiento': st.session_state.get(f'pa_fecha_seg_nueva_{i}', ''),
                    'estado': st.session_state.get(f'pa_estado_nueva_{i}', ''),
                    'costo_seguimiento': convertir_a_numero(st.session_state.get(f'pa_costo_seg_nueva_{i}', 0)),
                    'evidencia': st.session_state.get(f'pa_evidencia_nueva_{i}', '')
                }
        
        # 3. Asignar filas a cada causa y crear matriz de asignación
        fila_actual = 53  # Plan de Acción comienza en fila 53
        causa_filas = {}  # {causa_nombre: [fila_inicio, fila_fin]}
        actividad_filas = {}  # {actividad_idx: [fila_inicio, fila_fin]}
        
        # Calcular cuántas filas necesita cada causa
        causas_con_actividades = {}
        for actividad_idx, datos in actividades_map.items():
            for causa in datos['causas']:
                if causa not in causas_con_actividades:
                    causas_con_actividades[causa] = []
                causas_con_actividades[causa].append(actividad_idx)
        
        # Asignar filas a cada causa
        for causa, actividades_indices in causas_con_actividades.items():
            num_filas_causa = len(actividades_indices)
            causa_filas[causa] = [fila_actual, fila_actual + num_filas_causa - 1]
            
            # Asignar filas a cada actividad de esta causa
            for i, actividad_idx in enumerate(actividades_indices):
                fila_actividad = fila_actual + i
                if actividad_idx not in actividad_filas:
                    actividad_filas[actividad_idx] = []
                actividad_filas[actividad_idx].append(fila_actividad)
            
            fila_actual += num_filas_causa
        
        # 4. Escribir causas y crear combinaciones verticales para causas
        for causa, (fila_inicio, fila_fin) in causa_filas.items():
            causa_text = causas_map.get(causa, '')
            if causa_text:
                # Escribir la causa en la primera fila
                escribir_celda_segura(f'A{fila_inicio}', causa_text)
                
                # Combinar celdas A-D verticalmente si hay múltiples filas
                if fila_inicio < fila_fin:
                    try:
                        sheet.merge_cells(f'A{fila_inicio}:D{fila_fin}')
                        print(f"DEBUG: Combinando causa '{causa}': A{fila_inicio}:D{fila_fin}")
                    except Exception as e:
                        print(f"DEBUG: Error combinando causa '{causa}': {e}")
                else:
                    # Solo una fila, combinar horizontalmente
                    try:
                        sheet.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
                        print(f"DEBUG: Combinando causa '{causa}' (1 fila): A{fila_inicio}:D{fila_inicio}")
                    except Exception as e:
                        print(f"DEBUG: Error combinando causa '{causa}' (1 fila): {e}")
        
        # 5. Escribir actividades y crear combinaciones para actividades
        for actividad_idx, datos in actividades_map.items():
            filas_actividad = actividad_filas.get(actividad_idx, [])
            
            if filas_actividad:
                # Escribir la actividad en la primera fila donde aparece
                primera_fila = min(filas_actividad)
                ultima_fila = max(filas_actividad)
                
                # Escribir datos de la actividad
                escribir_celda_segura(f'E{primera_fila}', datos['texto'])
                
                # Combinar E-I verticalmente si la actividad está en múltiples filas
                if primera_fila < ultima_fila:
                    try:
                        sheet.merge_cells(f'E{primera_fila}:I{ultima_fila}')
                        print(f"DEBUG: Combinando actividad {actividad_idx}: E{primera_fila}:I{ultima_fila}")
                    except Exception as e:
                        print(f"DEBUG: Error combinando actividad {actividad_idx}: {e}")
                else:
                    # Solo una fila, combinar horizontalmente
                    try:
                        sheet.merge_cells(f'E{primera_fila}:I{primera_fila}')
                        print(f"DEBUG: Combinando actividad {actividad_idx} (1 fila): E{primera_fila}:I{primera_fila}")
                    except Exception as e:
                        print(f"DEBUG: Error combinando actividad {actividad_idx} (1 fila): {e}")
                
                # Escribir otros campos en todas las filas donde aparece la actividad
                for fila in filas_actividad:
                    if datos['responsable_ej']:
                        escribir_celda_segura(f'J{fila}', datos['responsable_ej'])
                        try:
                            sheet.merge_cells(f'J{fila}:K{fila}')
                        except: pass
                    
                    if datos['tiempo'] > 0:
                        escribir_celda_segura(f'L{fila}', datos['tiempo'])
                        try:
                            sheet.merge_cells(f'L{fila}:M{fila}')
                        except: pass
                    
                    if datos['costo'] > 0:
                        escribir_celda_segura(f'N{fila}', datos['costo'])
                        try:
                            sheet.merge_cells(f'N{fila}:O{fila}')
                        except: pass
                    
                    if datos['fecha_inicio']:
                        escribir_celda_segura(f'P{fila}', formatear_fecha(datos['fecha_inicio']))
                        try:
                            sheet.merge_cells(f'P{fila}:Q{fila}')
                        except: pass
                    
                    if datos['fecha_fin']:
                        escribir_celda_segura(f'R{fila}', formatear_fecha(datos['fecha_fin']))
                        try:
                            sheet.merge_cells(f'R{fila}:S{fila}')
                        except: pass
                    
                    if datos['responsable_seg']:
                        escribir_celda_segura(f'T{fila}', datos['responsable_seg'])
                        try:
                            sheet.merge_cells(f'T{fila}:U{fila}')
                        except: pass
                    
                    if datos['fecha_seguimiento']:
                        escribir_celda_segura(f'V{fila}', formatear_fecha(datos['fecha_seguimiento']))
                    
                    if datos['estado'] and datos['estado'] != "":
                        escribir_celda_segura(f'W{fila}', datos['estado'])
                    
                    if datos['costo_seguimiento'] > 0:
                        escribir_celda_segura(f'X{fila}', datos['costo_seguimiento'])
                        try:
                            sheet.merge_cells(f'X{fila}:Y{fila}')
                        except: pass
                    
                    if datos['evidencia']:
                        escribir_celda_segura(f'Z{fila}', datos['evidencia'])
        
        # Ocultar filas vacías restantes
        if fila_actual <= 72:
            for fila in range(fila_actual, 73):
                sheet.row_dimensions[fila].hidden = True
        
        # QUINTA SECCIÓN: COSTOS ASOCIADOS (Filas 86 y 89)
        costo_correccion = convertir_a_numero(st.session_state.get('costo_correccion', 0))
        costo_reputacional = convertir_a_numero(st.session_state.get('costo_reputacional', 0))
        costo_acciones = convertir_a_numero(st.session_state.get('costo_acciones', 0))
        multas_sanciones = convertir_a_numero(st.session_state.get('multas_sanciones', 0))
        costo_seguimiento_final = convertir_a_numero(st.session_state.get('costo_seguimiento', 0))
        otros_costos_internos = convertir_a_numero(st.session_state.get('otros_costos_internos', 0))
        descuentos_cliente = convertir_a_numero(st.session_state.get('descuentos_cliente', 0))
        otros_costos = convertir_a_numero(st.session_state.get('otros_costos', 0))
        
        # Costos en la fila 86
        if costo_correccion > 0:
            escribir_celda_segura('A86', costo_correccion)
        if costo_acciones > 0:
            escribir_celda_segura('H86', costo_acciones)
        if costo_seguimiento_final > 0:
            escribir_celda_segura('O86', costo_seguimiento_final)
        if descuentos_cliente > 0:
            escribir_celda_segura('V86', descuentos_cliente)
        
        # Costos en la fila 89
        if costo_reputacional > 0:
            escribir_celda_segura('A89', costo_reputacional)
        if multas_sanciones > 0:
            escribir_celda_segura('H89', multas_sanciones)
        if otros_costos_internos > 0:
            escribir_celda_segura('O89', otros_costos_internos)
        if otros_costos > 0:
            escribir_celda_segura('V89', otros_costos)

        # Guardar en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error al generar Excel: {str(e)}")
        st.write(f"Error detallado: {type(e).__name__}: {str(e)}")
        return None

def generar_excel_acr(descripcion, area, causa_raiz, acciones, responsable, fecha_compromiso):
    """Genera el archivo Excel con los datos de la ACR"""
    try:
        data = {
            'Descripción del problema': [descripcion],
            'Área o proceso afectado': [area],
            'Análisis de causas (5 porqués)': [causa_raiz],
            'Acciones de mejora': [acciones],
            'Responsable': [responsable],
            'Fecha compromiso': [fecha_compromiso],
            'Fecha de generación': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            'Generado por': ['Solutions & Payroll - Sistema ACR']
        }
        
        df = pd.DataFrame(data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='ACR', index=False)
            
            # Ajustar ancho de columnas
            worksheet = writer.sheets['ACR']
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"❌ Error al generar el Excel: {str(e)}")
        return None

def procesar_archivos_excel(archivo_acr, archivo_maestro):
    """Procesa los archivos Excel para transferir datos del ACR al maestro"""
    try:
        df_acr = pd.read_excel(archivo_acr)
        df_maestro = pd.read_excel(archivo_maestro)
        
        df_resultado = pd.concat([df_maestro, df_acr], ignore_index=True)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_resultado.to_excel(writer, sheet_name='Historial_ACR', index=False)
            
            worksheet = writer.sheets['Historial_ACR']
            for idx, col in enumerate(df_resultado.columns):
                max_length = max(df_resultado[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"❌ Error al procesar los archivos: {str(e)}")
        return None

if __name__ == "__main__":
    main()