import openpyxl

wb = openpyxl.load_workbook('Formato ACR - limpio.xlsx')
ws = wb.active

print("=" * 80)
print("ANÁLISIS DEL FORMATO EXCEL ACR")
print("=" * 80)

# Secciones importantes
secciones = {
    "INFORMACIÓN GENERAL": (3, 10),
    "CORRECCIÓN": (11, 16),
    "IDENTIFICACIÓN DE CAUSAS": (17, 32),
    "PLAN DE ACCIÓN": (30, 60),
    "COSTOS": (60, 80)
}

for nombre, (inicio, fin) in secciones.items():
    print(f"\n{'='*80}")
    print(f"{nombre} (Filas {inicio}-{fin})")
    print(f"{'='*80}")
    
    for fila in range(inicio, min(fin, inicio + 10)):
        valores = []
        for col in range(1, 15):  # A-N
            val = ws.cell(fila, col).value
            if val:
                letra = chr(64 + col)
                valores.append(f"{letra}{fila}:{val[:50] if len(str(val)) > 50 else val}")
        
        if valores:
            print(f"\nFila {fila}:")
            for v in valores:
                print(f"  {v}")

print(f"\n{'='*80}")
print("CELDAS COMBINADAS EN PLAN DE ACCIÓN (filas 30-60)")
print(f"{'='*80}")

merged = list(ws.merged_cells.ranges)
for r in merged:
    if r.min_row >= 30 and r.max_row <= 60:
        print(f"  {str(r)}")

print(f"\n{'='*80}")
print("FIN DEL ANÁLISIS")
print(f"{'='*80}")
