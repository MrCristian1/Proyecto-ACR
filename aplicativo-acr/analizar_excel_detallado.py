import openpyxl

wb = openpyxl.load_workbook('Formato ACR - limpio.xlsx')
ws = wb.active

print("=" * 100)
print("MAPEO DETALLADO DE CELDAS - FORMATO ACR")
print("=" * 100)

# Revisar todas las filas con contenido
print("\n" + "="*100)
print("CONTENIDO DE FILAS IMPORTANTES")
print("="*100)

filas_importantes = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 17, 18, 19, 20, 30, 31, 32, 33, 34, 50, 51, 60, 61, 62, 70, 71, 72]

for fila in filas_importantes:
    contenido = []
    for col in range(1, 29):  # A-AB (columnas 1-28)
        val = ws.cell(fila, col).value
        if val and str(val).strip():
            letra = openpyxl.utils.get_column_letter(col)
            contenido.append(f"{letra}:{val}")
    
    if contenido:
        print(f"\nFila {fila}:")
        for item in contenido:
            print(f"  {item}")

print("\n" + "="*100)
print("SECCIÓN: COSTOS ASOCIADOS (buscando en filas 60-80)")
print("="*100)

for fila in range(60, 80):
    contenido = []
    for col in range(1, 29):
        val = ws.cell(fila, col).value
        if val and str(val).strip():
            letra = openpyxl.utils.get_column_letter(col)
            contenido.append(f"{letra}{fila}:{val}")
    
    if contenido:
        print(f"\nFila {fila}: {', '.join(contenido)}")

print("\n" + "="*100)
print("ANÁLISIS DE CELDAS PARA PLAN DE ACCIÓN")
print("="*100)
print("\nBuscando encabezados de columnas en fila 33 y alrededores...")

for fila in [30, 31, 32, 33, 49, 50, 51]:
    print(f"\nFila {fila}:")
    for col in range(1, 29):
        val = ws.cell(fila, col).value
        if val:
            letra = openpyxl.utils.get_column_letter(col)
            print(f"  {letra}{fila}: {val}")
